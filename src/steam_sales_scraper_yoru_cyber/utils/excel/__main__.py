import logging
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import List

from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from steam_sales_scraper_yoru_cyber.bot.__main__ import Product

purple_bg = PatternFill(start_color="3e2f5b", fill_type="solid")
green_bg = PatternFill(start_color="71b340", fill_type="solid")
base_font = Font(size=11, bold=True, color="000000", name="Courier New")
base_alignment = Alignment(horizontal="center", vertical="center")


def write_csv(product_list: List[Product], fname: str | Path):
    count_lines = 0
    for product in product_list:
        logging.info(f"Saving data in CSV file line: {count_lines}")
        with open(f"{fname}-{datetime.now().strftime('%d-%m-%Y-%H:%M:%S')}.txt", "a") as f:
            f.writelines(f"{product.name},{product.price}\n")
            count_lines += 1


def write_names(ws, product_list: List):
    for row in ws.iter_cols(min_row=2, max_row=ws.max_row + len(product_list), min_col=1, max_col=1):
        for cell, product in zip(row, product_list):
            cell.alignment = base_alignment
            cell.value = product.name


def write_prices(ws, product_list: List):
    for row in ws.iter_cols(min_row=2, max_row=ws.max_row + len(product_list),
                            min_col=2, max_col=2):
        for cell, product in zip(row, product_list):
            cell.value = product.price


def write_headers(ws: Worksheet, bg: PatternFill, name: str, cell: str):
    ws[cell].value = name
    ws[cell].font = base_font
    ws[cell].fill = bg
    ws[cell].alignment = base_alignment


def to_excel(products_list: List, wb: Workbook):
    worksheets = list()
    ws_all_items = wb.create_sheet("All Items", 0)
    ws_bestsellers = wb.create_sheet("Bestsellers", 1)
    ws_new_trending = wb.create_sheet("New & Trending", 2)
    write_headers(ws_all_items, purple_bg, "Name", "A1")
    write_headers(ws_all_items, green_bg, "Name", "B1")
    write_headers(ws_bestsellers, purple_bg, "Name", "A1")
    write_headers(ws_bestsellers, green_bg, "Name", "B1")
    write_headers(ws_new_trending, purple_bg, "Name", "A1")
    write_headers(ws_new_trending, green_bg, "Name", "B1")
    worksheets.append(ws_all_items)
    worksheets.append(ws_bestsellers)
    worksheets.append(ws_new_trending)
    for product_list, ws in zip(products_list, worksheets):
        write_names(ws, product_list)
        write_prices(ws, product_list)
    return wb
