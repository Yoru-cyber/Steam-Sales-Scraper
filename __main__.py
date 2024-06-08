import time
import sys
import logging
import os
import concurrent.futures
from pathlib import Path
from dataclasses import dataclass
from datetime import datetime
from typing import List


from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from playwright.sync_api import sync_playwright

today_date = datetime.now().strftime('%d-%m-%Y-%H:%M:%S')
if not os.path.exists("./logs"):
    os.mkdir("logs")
if not os.path.exists("./data"):
    os.mkdir("data")
logging.basicConfig(level=logging.INFO, filename=Path(f"./logs/app-{today_date}.log"),
                    filemode="w",
                    format="%(asctime)s %(threadName)s  %(levelname)s - %(message)s")
logging.info("Process starting")



@dataclass
class Product:
    name: str
    price: str


purple_bg = PatternFill(start_color="3e2f5b", fill_type="solid")
green_bg = PatternFill(start_color="71b340", fill_type="solid")
base_font = Font(size=11, bold=True, color="000000", name="Courier New")
base_alignment = Alignment(horizontal="center", vertical="center")

logging.getLogger().addHandler(logging.StreamHandler(sys.stdout))


def scrape_data(url: str) -> List[Product]:
    logging.info("Getting data")
    product_list: List[Product] = list()
    with sync_playwright() as playwright:
        chrome = playwright.chromium
        browser = chrome.launch(headless=False)
        page = browser.new_page()
        page.goto(url, timeout=10000, wait_until="domcontentloaded")
        time.sleep(5)
        logging.info(f"DOM content loaded from {url}")
        # i = 4
        # while i > 0:
        #     page.evaluate("() => window.scrollBy(0, document.body.scrollHeight)")
        #     page.wait_for_selector("._3d9cKhzXJMPBYzFkB_IaRp.Focusable").click()
        #     time.sleep(2)
        #     i -= 1
        page.evaluate("() => window.scrollBy(0, document.body.scrollHeight)")
        page.wait_for_selector(".y9MSdld4zZCuoQpRVDgMm")
        div_locators = page.query_selector_all(".y9MSdld4zZCuoQpRVDgMm")
        for element in div_locators:
            anchor_locator = element.query_selector("a")
            if anchor_locator:
                url = anchor_locator.get_attribute("href")
                if url is not None:
                    new_page = browser.new_page()
                    new_page.goto(url, timeout=10000, wait_until="domcontentloaded")
                    logging.info(f"DOM content load from {url}")
                    if new_page.query_selector(".age_gate") is not None:
                        new_page.wait_for_load_state("domcontentloaded", timeout=10000)
                        time.sleep(5)
                        new_page.query_selector("select[id='ageYear']").select_option(
                            "2000"
                        )
                        new_page.query_selector("a[id='view_product_page_btn']").click()
                        new_page.wait_for_load_state("domcontentloaded", timeout=10000)
                        time.sleep(10)
                        logging.info(f"DOM content load from {url}")
                        game = Product(
                            name=new_page.wait_for_selector(
                                ".apphub_AppName", timeout=10000
                            ).inner_text(),
                            price=new_page.wait_for_selector(
                                ".discount_final_price", timeout=10000
                            ).inner_text(),
                        )
                        logging.info(f"Product Name: {game.name}, Product Price: {game.price}")
                        product_list.append(game)
                        new_page.close()
                    else:
                        new_page.wait_for_load_state("domcontentloaded", timeout=10000)
                        time.sleep(10)
                        game = Product(
                            name=new_page.wait_for_selector(
                                ".apphub_AppName", timeout=10000
                            ).inner_text(),
                            price=new_page.wait_for_selector(
                                ".discount_final_price", timeout=10000
                            ).inner_text(),
                        )
                        logging.info(f"Product Name: {game.name}, Product Price: {game.price}")
                        product_list.append(game)
                        new_page.close()
                else:
                    pass
        browser.close()
        return product_list


def write_csv(product_list: List[Product], fname: str | Path):
    count_lines = 0
    for product in product_list:
        logging.info(f"Saving data in CSV file line: {count_lines}")
        with open(f"{fname}-{datetime.now().strftime('%d-%m-%Y-%H:%M:%S')}.txt", "a") as f:
            f.writelines(f"{product.name},{product.price}\n")
            count_lines += 1


def write_names(ws, product_list: List):
    for row in ws.iter_cols(min_row=2, max_row=ws.max_row + len(product_list), min_col=1, max_col=1):
        for cell, product in zip(row, product_list):
            cell.alignment = base_alignment
            cell.value = product.name


def write_prices(ws, product_list: List):
    for row in ws.iter_cols(min_row=2, max_row=ws.max_row + len(product_list),
                            min_col=2, max_col=2):
        for cell, product in zip(row, product_list):
            cell.value = product.price


def to_excel(products_list: List, file: Path | str):
    wb = Workbook()
    worksheets = list()
    ws_all_items = wb.create_sheet("All Items", 0)
    ws_bestsellers = wb.create_sheet("Bestsellers", 1)
    ws_new_trending = wb.create_sheet("New & Trending", 2)
    ws_all_items['A1'].value = "Name"
    ws_bestsellers['A1'].value = "Name"
    ws_new_trending['A1'].value = "Name"
    ws_all_items['B1'].value = "Price"
    ws_bestsellers['B1'].value = "Price"
    ws_new_trending['B1'].value = "Price"
    ws_all_items['A1'].value = "Name"
    ws_all_items['A1'].font = base_font
    ws_all_items['A1'].fill = purple_bg
    ws_all_items['A1'].alignment = base_alignment
    ws_all_items['B1'].value = "Price"
    ws_all_items['B1'].font = base_font
    ws_all_items['B1'].fill = green_bg
    ws_all_items['B1'].alignment = base_alignment
    ws_bestsellers['A1'].value = "Name"
    ws_bestsellers['A1'].font = base_font
    ws_bestsellers['A1'].fill = purple_bg
    ws_bestsellers['B1'].value = "Price"
    ws_bestsellers['B1'].font = base_font
    ws_bestsellers['B1'].alignment = base_alignment
    ws_bestsellers.fill = green_bg
    ws_new_trending['A1'].value = "Name"
    ws_new_trending['A1'].font = base_font
    ws_new_trending['A1'].fill = purple_bg
    ws_new_trending['A1'].alignment = base_alignment
    ws_new_trending['B1'].value = "Price"
    ws_new_trending['B1'].font = base_font
    ws_new_trending.fill = green_bg
    ws_new_trending['B1'].alignment = base_alignment
    worksheets.append(ws_all_items)
    worksheets.append(ws_bestsellers)
    worksheets.append(ws_new_trending)
    for product_list, ws in zip(products_list, worksheets):
        write_names(ws, product_list)
        write_prices(ws, product_list)
    wb.save(file)


def main():
    start_time = time.time()
    urls = ["https://store.steampowered.com/specials/?flavor=contenthub_topsellers",
            "https://store.steampowered.com/specials/?flavor=contenthub_newandtrending",
            "https://store.steampowered.com/specials/"]
    logging.info("App started")

    results = list()
    with concurrent.futures.ThreadPoolExecutor(max_workers=3) as executor:
        best_sellers = executor.submit(scrape_data, urls[0])
        new_trending = executor.submit(scrape_data, urls[1])
        all_items = executor.submit(scrape_data, urls[2])
        best_sellers_result = best_sellers.result()
        new_trending_result = new_trending.result()
        all_items_result = all_items.result()
        results.append(all_items_result)
        results.append(best_sellers_result)
        results.append(new_trending_result)
    fnames = [Path("./data/All_items"),
              Path("./data/Best_Sellers"),
              Path("./data/New&Trending")]
    write_csv(results[0], fnames[0])
    write_csv(results[1], fnames[1])
    write_csv(results[2], fnames[2])
    to_excel(results, Path(f"./data/app-{today_date}.xlsx"))
    logging.info(f"Finished in {time.time() - start_time}")


if __name__ == "__main__":
    main()
